from project import pay_raise, xl, pay_cut
from openpyxl import  load_workbook
import pytest



def test_pay_raise():
    details = [
        {"Name": "Tej", "Designation": "Manager", "Pay": "$45"},
        {"Name": "John", "Designation": "Programmer", "Pay": "$25"},
        {"Name": "Andrew", "Designation": "Hr", "Pay": "$30"}
    ]
    xl(details,'test_employee.xlsx')
    to_raise='Tej'
    r=10
    pay_raise(to_raise,r,'test_employee.xlsx')
    workbook=load_workbook('test_employee.xlsx')
    worksheet=workbook.active
    cell=worksheet['D2'].value
    assert cell=='$55'
    assert pay_raise('some_random_name',5,'test_employee.xlsx')=='No one named some_random_name was found in the database.'


def test_pay_cut():
    details = [
        {"Name": "Tej", "Designation": "Manager", "Pay": "$45"},
        {"Name": "John", "Designation": "Programmer", "Pay": "$25"},
        {"Name": "Andrew", "Designation": "Hr", "Pay": "$30"}
    ]
    xl(details, 'test_employee.xlsx')
    to_cut = 'Tej'
    c = 5
    pay_cut(to_cut, c, 'test_employee.xlsx')
    workbook = load_workbook('test_employee.xlsx')
    worksheet = workbook.active
    cell = worksheet['D2'].value
    assert cell == "$40"
    assert pay_cut('Tej',50,'test_employee.xlsx')=="You can't cut all of the pay"



    result = pay_cut('some_random_name', c, 'test_employee.xlsx')
    assert result == 'No one named some_random_name was found in the database.'
def test_xl():
    details = [
        {"Name": "Tej", "Designation": "Manager", "Pay": "$45"},
        {"Name": "John", "Designation": "Programmer", "Pay": "$25"},
        {"Name": "Andrew", "Designation": "Hr", "Pay": "$30"}
    ]
    xl(details,'test_employee.xlsx')
    workbook=load_workbook('test_employee.xlsx')
    worksheet=workbook.active

    assert worksheet['A1'].value=='S.no'
    assert worksheet['A2'].value==1
    assert worksheet['A3'].value==2
    assert worksheet['A4'].value==3
    assert worksheet['B1'].value=='Name'
    assert worksheet['B2'].value=='Tej'
    assert worksheet['B3'].value=='John'
    assert worksheet['B4'].value=='Andrew'
    assert worksheet['C1'].value=='Designation'
    assert worksheet['C2'].value=='Manager'
    assert worksheet['C3'].value=='Programmer'
    assert worksheet['C4'].value=='Hr'
    assert worksheet['D1'].value=='Pay'
    assert worksheet['D2'].value=='$45'
    assert worksheet['D3'].value=='$25'
    assert worksheet['D4'].value=='$30'





