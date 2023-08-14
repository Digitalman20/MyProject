from openpyxl import Workbook,load_workbook
import smtplib
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
from cryptography.fernet import Fernet


'''
The base principle of this code is, it takes in multiple inputs(details of the employee, such as 'Name', 'Designation' and 'Hourly_pay')
and creates an excel sheet out of the given input. There are functions in the code that deal with hiring an employee(i.e adding them to the excel sheet)
, firing an employee(i.e, removing them from the sheet and adjusting the serial number accordingly) and making changes to the pay.
There is also a function that mails the  excel sheet to the inputed mail I'd.
'''

#The below list is used to store the input from the function(basic_details).
details=[]
"""Basically the below function takes in user input and appends that input the above list in the form of multiple dictionaries.
   The input for designation can only be the keys from the valid_post dictionary. The value of each key in the dictionary is a tuple containing the
   pay range. The pay can only be a value in the range. You can only add one person per iteration, there is a prompt to add more people."""
def basic_details():
    #The variable below starts the loop.
    add_more='y'
    #The dictionary below contains the designations and the pay range.
    valid_post={
        "Manager":(40,50),
        "Assistant":(15,20),
        "Programmer":(20,30),
        "Hr":(20,30),
        "Clerk":(15,20)}

    print("The posts and hourly pay range for the employee are as follow:")
    print(valid_post)

    #This while loop
    while add_more=='y':
        name: str=str(input('Name:')).title().strip()
        post=''
        while post not in valid_post:
            post=(input('Designation:')).title().strip()
        min_pay,max_pay=valid_post[post]
        pay=int()
        while pay>max_pay or pay<min_pay:
            pay=(input('Hourly pay:')).strip()
            pay=int(pay)

        details.append({'Name':name,'Designation':post,'Pay':f'${pay}'})
        add_more=input('Would you like to add more people? y for yes;').lower()
    return details

"""The function xl takes in two parameter, a list of dictionaries and the name of filein this case, and creates an excel sheet from the text
   in the parameter. This function can be used to overwrite the existing excel sheet or to create a fresh one."""
def xl(list,file):
    workbook=Workbook()
    worksheet=workbook.active
    position=1
    #The four lines below are used to write the headers in the excel sheet.
    worksheet.cell(row=1,column=1,value='S.no')
    worksheet.cell(row=1,column=2,value='Name')
    worksheet.cell(row=1,column=3,value='Designation')
    worksheet.cell(row=1,column=4,value='Pay')


    #The 'for loop' goes through the list and writes each dictionary in new rows accordingly.
    for position,text in enumerate(list):

        worksheet.append([position+1,text['Name'],text['Designation'],text['Pay']])

    print('Saving file ....')
    workbook.save(file)
    print('Saved.')

#Function hire also takes in one parameter, and appends the  text in the parameter to an existing excel sheet.
def hire(list,file):
    wb=load_workbook(file)
    ws=wb.active
    row=ws.max_row
    #The 'for loop' goes through the list and writes each dictionary in new rows accordingly.
    for position,text in enumerate(list):
        ws.append([row,text['Name'],text['Designation'],text['Pay']])
        row+=1


    wb.save(file)


#Function fire removes a row in an excel sheet based on the input given.
def fire():
    to_fire=input('Fire who?').title()
    workbook=load_workbook('Employee.xlsx')

    worksheet=workbook.active
    mrow=worksheet.max_row
    #The 'for loop' below goes through every row in column['B']
    for row in worksheet.iter_cols(2):
        #The 'for loop' below goes through every cell in each row in column['B'].
        for cell in row:
            # The 'if statement' below checks if the input matches with the value in each cell and removes the cell if so.
            if cell.value==to_fire or cell.value==None :
                worksheet.delete_rows(cell.row)
                cl='B'+str(mrow)
                '''The below statement is included to avoid a bug. When a row is deleted from the center of the sheet an empty row is left at
                the bottom of the file.When hire runs after fire a line and a serial number is skipped.'''
                if worksheet[cl].value==None:
                    worksheet.delete_rows(mrow)

            else:
                pass
    rown=1
    #The following 'for loop' rewrites the serial number column in ascending order.
    for row in range(2,mrow):
        worksheet.cell(row=row,column=1,value=rown)
        rown+=1
    workbook.save('Employee.xlsx')

#pay_raise funtion is used to increse an employee's hourly pay.
def pay_raise(to_raise,r,file):
    workbook=load_workbook(file)
    worksheet=workbook.active
    mrow=worksheet.max_row
    updated_pay=None
    '''This 'nested for loop' with the 'if else statement' goes through each cell in column['B'] and matches the input(to_raise).
       If the input matches in a row, it takes the value in the pay column of that row and adds the input(r) to it.'''
    for row in range(2,mrow+1):
        for column in range(1,2):
            index=chr(65+column)
            cl=index+str(row)
            if worksheet[cl].value==to_raise:
                new_index='D'+str(row)
                pay=worksheet[new_index].value
                pay=int(pay.strip('$'))
                updated_pay='$'+str(pay+r)
                worksheet.cell(row=row,column=4,value=updated_pay)
            else:
                pass
    if updated_pay==None:
        print(f'No one named {to_raise} was found in the database.')
        updated_pay=f'No one named {to_raise} was found in the database.'
    print('Saving ...')
    workbook.save(file)
    return updated_pay

def pay_cut(to_cut, c, file):

    updated_pay=None
    workbook=load_workbook(file)
    worksheet=workbook.active
    mrow=worksheet.max_row
    '''This 'nested for loop' with the 'if else statement' goes through each cell in column['B'] and matches the input(to_cut).
       If the input matches in a row, it takes the value in the pay column of that row and cuts the input(c) to it.The while loop
       denies extra pay being cut(i.e,more than the current pay).'''
    for row in range(2,mrow+1):
        for column in range(1,2):
            index=chr(65+column)
            cl=index+str(row)
            if worksheet[cl].value==to_cut:
                new_index='D'+str(row)
                pay=worksheet[new_index].value
                pay=int(pay.strip('$'))


                result=pay-c
                if result>0:
                    updated_pay='$'+str(result)
                    worksheet.cell(row=row,column=4,value=updated_pay)
                    workbook.save(file)
                    print('Saving ...')
                    return updated_pay
                elif result<=0:
                     statement="You can't cut all of the pay"
                     print(statement)
                     return statement

    if updated_pay==None :
        print(f'No one named {to_cut} was found in the database.')
        statement=f'No one named {to_cut} was found in the database.'
        return statement


#The function mail just sends the excel sheet to the inputed mail.
def mail():
    #The below statement is added to decrypt the encrypted password file.

    with open('key.key','rb') as file:
        key=file.read()

    with open('password.key','rb') as file:
        enc=file.read()
    f=Fernet(key)
    password=f.decrypt(enc).decode()

    smtp_obj=smtplib.SMTP('smtp.gmail.com',587)
    smtp_obj.ehlo()
    smtp_obj.starttls()
    mail=MIMEMultipart()

    from_address='pythonprojectemp@gmail.com'


    smtp_obj.login(from_address,password)

    to_address=input('Send to:')

    mail['From']=from_address
    mail['To']=to_address
    mail['Subject']='Employee Details'
    body=MIMEText('This enclosed file is the latest employee details.','plain')
    mail.attach(body)

    to_attach = MIMEBase('application', "octet-stream")
    to_attach.set_payload(open("Employee.xlsx", "rb").read())
    encoders.encode_base64(to_attach)
    to_attach.add_header('Content-Disposition', 'attachment; filename="Employee.xlsx"')
    mail.attach(to_attach)
    smtp_obj.send_message(mail,from_addr=from_address,to_addrs=[to_address])

def mail_choice():
    #The below ensures if the mail is to be sent or not and executes the mailing process.

    mail_choice=input('Would you like to email the updated file? "y" for yes, anything else for no; ').lower()

    if mail_choice=='y':
        mail()
    else:
        pass


def main():
    do_it_again='y'
    #The below 'while loop' is used to repeat program based on a certain input.
    while do_it_again=='y':
        #The 'print function' is used to print the menu.
        print("Press '1' to write a new excel file")
        print("Press '2' to 'hire' one or more employee(s)")
        print("Press '3' to 'fire' an existing employee")
        print("Press '4' to make changes to the pay of any employee")
        print("Press '5' to email the excel file")
        print("Press '8' to exit")
        #The variable 'mail_choice' is included to prompt to mail in each option of the menu.
        valid_options=['1','2','3','4','5','8']
        option=''
        while option not in valid_options:
            option=input('What would you like to do? ')
        #This runs menu option 1.
        if option=='1':
            basic_details()
            xl(details,'Employee.xlsx')
            mail_choice()
            #If the input of the below is 'y', the program runs again.
            do_it_again=input('Would you like to do anything else? "y" for yes, anything else to exit; ').lower()

        #The below runs menu option 2
        elif option=='2':
            basic_details()
            hire(details,'Employee.xlsx')
            mail_choice()
            #If the input of the below is 'y', the program runs again.
            do_it_again=input('Would you like to do anything else? "y" for yes, anything else to exit; ').lower()
        #The below runs menu option 3
        elif option=='3':
            fire()
            mail_choice()
            #If the input of the below is 'y', the program runs again.
            do_it_again=input('Would you like to do anything else? "y" for yes, anything else to exit; ').lower()

        #The below runs menu option 4
        elif option =='4':
            valid_choice=['6','7']
            choice=''
            #The 'while loop' only takes an input from valid_choice and makes sures to run the correct function for each choice.
            while choice not in valid_choice:
                choice=input('Would you like to pay raise or pay cut? 6 for pay raise and 7 for pay cut; ')
            if choice=='6':
                to_raise=input('Whom do you want to raise?').title()
                while True:
                    try:
                       r=int(input('How much would you like to raise?'))
                       break
                    except ValueError:
                        print('You can only use integers to specify the amount you wanna cut.')
                pay_raise(to_raise,r,'Employee.xlsx')
            elif choice=='7':
                c=None
                to_cut=input('Whose pay would you like to cut?').title()
                while True:
                    while True:
                        try:
                           c=int(input('How much would you like to cut? '))
                           break
                        except ValueError:
                            print('You can only use integers to specify the amount you wanna cut.')
                    result=pay_cut(to_cut, c, 'Employee.xlsx')
                    if result[1:].isdigit():
                        if int(result.strip('$'))>0:
                            break
                        else:
                            pass
                    else:
                        pass
            mail_choice()
            #If the input of the below is 'y', the program runs again.
            do_it_again=input('Would you like to do anything else? "y" for yes, anything else to exit; ').lower()
        #The below runs menu option 5
        elif option=='5':
            mail()
            #If the input of the below is 'y', the program runs again.
            do_it_again=input('Would you like to do anything else? "y" for yes, anything else to exit; ').lower()
        elif option=='8':
            break
if __name__=='__main__':
    main()



